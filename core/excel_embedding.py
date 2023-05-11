import os
from core.embedding import create_embedding
from core.vector_db import Storage

import openpyxl


def read_excel(excel_file: str):
    bk = openpyxl.load_workbook(excel_file)
    sheet = bk.active
    maxrow = sheet.max_row
    minrow = sheet.min_row+1  # skip table header
    for i in range(minrow, maxrow + 1):
        q_value = sheet.cell(i, 1).value
        a_value = sheet.cell(i, 2).value
        yield q_value, a_value  # read QA data


def excel_to_db(excel_file: str) -> None:
    storage = Storage()
    id_count = 1
    for q_value, a_value in read_excel(excel_file):
        try:
            _, vector = create_embedding(str(q_value))
        except Exception as exce:
            print(str(exce))
            input("wait for command to retry")
            _, vector = create_embedding(str(q_value))
        storage.add(id_count, a_value, 0, vector)
        id_count += 1
        print(f"> 完成插入text: [{a_value[0:10]}], embedding: {vector[0:3]}")
    del storage


if __name__ == '__main__':
    excel_to_db('/Users/abcd/Desktop/IM知识库.xlsx')
